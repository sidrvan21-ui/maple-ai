"""Turn whatever a company drops into text Maple can index."""

from __future__ import annotations

import io
import json
import re
import zipfile
from email import message_from_bytes
from pathlib import Path
from xml.etree import ElementTree

from app.onboard import slug_product, stage_from_path
from app.rag.admit import STAGE_FOLDERS, repo_root

MAX_TOTAL_BYTES = 40 * 1024 * 1024
MAX_FILES = 80
BLOCKED = {
    ".exe",
    ".bat",
    ".cmd",
    ".com",
    ".scr",
    ".dll",
    ".msi",
    ".ps1",
    ".vbs",
    ".sh",
    ".bash",
    ".zsh",
    ".app",
    ".dmg",
    ".iso",
}
SKIP_NAMES = {"__macosx", ".ds_store", "thumbs.db"}


def products_root() -> Path:
    path = repo_root() / "data" / "products"
    path.mkdir(parents=True, exist_ok=True)
    return path


def safe_name(name: str) -> str:
    base = Path(name.replace("\\", "/")).name
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return (cleaned or "file")[:120]


def is_blocked(name: str) -> bool:
    return Path(name).suffix.lower() in BLOCKED


def bytes_to_markdown(filename: str, data: bytes) -> tuple[str, str]:
    """Returns (markdown, how). how is extracted | stub."""
    suffix = Path(filename).suffix.lower()
    label = Path(filename).name
    try:
        if suffix in {".md", ".txt", ".text", ".log", ".csv", ".tsv"}:
            return _decode(data), "extracted"
        if suffix == ".json":
            parsed = json.loads(_decode(data))
            pretty = json.dumps(parsed, indent=2)[:20000]
            return "```json\n" + pretty + "\n```\n", "extracted"
        if suffix in {".html", ".htm"}:
            return _from_html(data), "extracted"
        if suffix == ".pdf":
            text = _from_pdf(data)
            if text:
                return text, "extracted"
            return _stub(label, "PDF had no extractable text (often a scan)."), "stub"
        if suffix == ".docx":
            return _from_docx(data), "extracted"
        if suffix in {".xlsx", ".xlsm"}:
            return _from_xlsx(data), "extracted"
        if suffix == ".pptx":
            text = _from_pptx(data)
            if text:
                return text, "extracted"
            return _stub(label, "PowerPoint had no extractable text."), "stub"
        if suffix in {".eml", ".msg"}:
            return _from_email(data, label), "extracted"
        if suffix == ".xml":
            return _decode(data)[:20000], "extracted"
    except Exception:
        return _stub(label, "Maple kept the original. Text extract failed."), "stub"
    return _stub(label, "Maple kept the original. This type is not auto-read yet."), "stub"


def _stub(label: str, reason: str) -> str:
    return f"# {label}\n\n{reason}\n\nOriginal filename: `{label}`\n"


def _decode(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _from_html(data: bytes) -> str:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(_decode(data), "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    return soup.get_text("\n", strip=True)


def _from_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages).strip()


def _from_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    chunks: list[str] = []
    for sheet in wb.worksheets:
        chunks.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                chunks.append(" | ".join(cells))
    wb.close()
    return "\n".join(chunks)


def _from_pptx(data: bytes) -> str:
    texts: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        slides = sorted(
            n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")
        )
        for name in slides:
            root = ElementTree.fromstring(zf.read(name))
            bits = [
                (node.text or "")
                for node in root.iter()
                if node.tag.endswith("}t") and node.text
            ]
            if bits:
                texts.append(" ".join(bits))
    return "\n\n".join(texts).strip()


def _from_email(data: bytes, label: str) -> str:
    msg = message_from_bytes(data)
    body = []
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True) or b""
                if isinstance(payload, bytes):
                    body.append(_decode(payload))
    else:
        payload = msg.get_payload(decode=True) or b""
        if isinstance(payload, bytes):
            body.append(_decode(payload))
    header = [
        f"# {label}",
        f"From: {msg.get('from', '')}",
        f"Subject: {msg.get('subject', '')}",
        "",
    ]
    return "\n".join(header + body)


def unpack_zip(raw: bytes) -> list[tuple[str, bytes, int | None]]:
    out: list[tuple[str, bytes, int | None]] = []
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile as exc:
        raise ValueError("not a zip file") from exc
    for info in zf.infolist():
        path = info.filename.replace("\\", "/")
        if ".." in Path(path).parts or path.startswith("/"):
            raise ValueError("zip has an unsafe path")
        base = Path(path).name.lower()
        if info.is_dir() or base in SKIP_NAMES or path.lower().startswith("__macosx/"):
            continue
        if is_blocked(path):
            continue
        out.append((path, zf.read(info), stage_from_path(path)))
    return out


def _unique(dir_path: Path, name: str) -> Path:
    candidate = dir_path / name
    if not candidate.exists():
        return candidate
    stem, suffix = Path(name).stem, Path(name).suffix
    n = 2
    while True:
        candidate = dir_path / f"{stem}_{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def list_products() -> list[str]:
    root = products_root()
    if not root.is_dir():
        return []
    names = []
    for path in sorted(root.iterdir()):
        if not path.is_dir() or path.name == "porter":
            continue
        if any((path / folder).is_dir() for folder in STAGE_FOLDERS.values()):
            names.append(path.name)
    return names


def resolve_product(product: str, existing: str) -> str:
    picked = (existing or "").strip()
    if picked:
        return picked
    return (product or "").strip()


def ingest(
    product: str,
    company: str,
    default_stage: int,
    uploads: list[tuple[str, bytes]],
) -> dict:
    slug = slug_product(product)
    if slug == "porter":
        raise ValueError("porter is Maple's sample. Use your product name.")
    if not uploads:
        raise ValueError("choose at least one file")
    items: list[tuple[str, bytes, int | None]] = []
    total = 0
    for name, data in uploads:
        total += len(data)
        if total > MAX_TOTAL_BYTES:
            raise ValueError("upload is larger than 40 MB")
        if name.lower().endswith(".zip"):
            items.extend(unpack_zip(data))
        else:
            if is_blocked(name):
                continue
            items.append((name, data, None))
    if not items:
        raise ValueError("no usable files (installers and scripts are blocked)")
    if len(items) > MAX_FILES:
        raise ValueError(f"too many files (max {MAX_FILES})")

    root = products_root() / slug
    orig = root / "_originals"
    orig.mkdir(parents=True, exist_ok=True)
    extracted = 0
    stubs = 0
    saved: list[str] = []
    rooms = {STAGE_FOLDERS[n]: 0 for n in STAGE_FOLDERS}

    for path, data, found_stage in items:
        stage = found_stage or default_stage
        folder = root / STAGE_FOLDERS[stage]
        folder.mkdir(parents=True, exist_ok=True)
        original = _unique(orig, safe_name(path))
        original.write_bytes(data)
        text, how = bytes_to_markdown(path, data)
        dest = _unique(folder, Path(safe_name(path)).stem + ".md")
        header = f"<!-- source: {Path(path).name} -->\n<!-- extract: {how} -->\n\n"
        dest.write_text(header + (text or _stub(Path(path).name, "empty")), encoding="utf-8")
        rooms[STAGE_FOLDERS[stage]] += 1
        saved.append(dest.relative_to(repo_root()).as_posix())
        if how == "extracted":
            extracted += 1
        else:
            stubs += 1

    return {
        "product": product.strip(),
        "company": company.strip(),
        "slug": slug,
        "files": len(saved),
        "extracted": extracted,
        "stubs": stubs,
        "rooms": rooms,
        "saved": saved[:80],
        "workbench": f"/p/{slug}",
    }
